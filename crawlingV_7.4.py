import requests 
from bs4 import BeautifulSoup
import datetime #날짜
import time # 일정시간 반복
import openpyxl
from openpyxl.styles import PatternFill
import sys
import os
# 텍스트 파일 생성 or 불러와서 urls 리스트에 추가
# urls 리스트 생성
urls = []

# Excel 파일생성
wb = openpyxl.Workbook()
sheet = wb.active

#색상지정
r_color = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
g_color = PatternFill(start_color='00e600', end_color='00e600', fill_type='solid')

def url_check():
    try:
        with open(os.path.dirname(os.path.realpath(__file__))+'/'+'url.txt', 'r') as file:
            for line in file:
                urls.append(line.strip('\n'))
        
    except:
        f = open(os.path.dirname(os.path.realpath(__file__))+'/'+"url.txt", 'w')
        f.close()
        print("생성된 url.txt 파일에 하나의 url을 입력하고 엔터를 반복한 후 다시 실행해주세요")
        sys.exit()

def make_column():
    for i in range(len(urls)):
        response = requests.get(urls[i])
        if response.status_code == 200:
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            name = soup.select_one('#name')
            name_text = name.get_text()
            print(name_text)
            # no_sheet_character = "\/*[]:?"

            # sheet이름에 특수문자 불가.
            #for x in range(len(no_sheet_character)):
            #    name_text = name_text.replace(no_sheet_character[x],"")
            #sheet 최대 문자열 길이 31개
            # if len(name_text) >30:
            #     name_text = name_text[:30]
            
            #sheet 이름에 제품명 붙이기
            # if i == 0 :
            #     wb['Sheet'].title = name_text
                
            # else:
            #     new_sheet = wb.create_sheet(name_text)
            sheet.cell(1, 1).value = "날짜/시간"
            sheet.cell(row = 1, column=i+2).value = name_text
        else:
            print("excel 구조 형성에서의 response status_code 오류")

    wb.save(os.path.dirname(os.path.realpath(__file__))+'/'+"newstock.xlsx")

def column_size():
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value   
    sheet.column_dimensions['A'].width = 25

def stock_check():
    row = 2
    # 재고유무 확인
    while True:

        for url_num in range(len(urls)):
            response = requests.get(urls[url_num])

            if response.status_code == 200 :
                #YYYY-MM-DD-HH-MM
                current = datetime.datetime.now()
                current_time = current.strftime('%Y-%m-%d %H:%M') #current.replace(microsecond=0)
                #time_list.append(str(current_time))
                html = response.text
                soup = BeautifulSoup(html, 'html.parser')
                stock_status = soup.select_one('#stock-status')
                text = stock_status.get_text()
                sheet.cell(row,1).value = current_time
                val = "품절" in text
                if val:
                    sheet.cell(row,url_num+2).value = "재고 X"
                    sheet.cell(row,url_num+2).fill = r_color
                else:
                    sheet.cell(row,url_num+2).value = "재고 O"
                    sheet.cell(row,url_num+2).fill = g_color
            else:
                print(response.status_code)
                print("비정상적인 접근")
        row += 1
        wb.save(os.path.dirname(os.path.realpath(__file__))+'/'+"newstock.xlsx")
        print('저장완료')
        time.sleep(5-len(urls))

def main():
    url_check()
    make_column()
    column_size()
    stock_check()

main()
