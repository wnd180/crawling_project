import requests
from bs4 import BeautifulSoup
import pandas as pd
import datetime #날짜
import time # 일정시간 반복
import openpyxl
from openpyxl.styles import PatternFill
import sys
import os
# 텍스트 파일 생성 or 불러와서 urls 리스트에 추가
# urls 리스트 생성
urls = []

try:
    with open(os.path.dirname(os.path.realpath(__file__))+'/'+'url.txt', 'r') as file:
        for line in file:
            urls.append(line.strip('\n'))
    
except:
    f = open(os.path.dirname(os.path.realpath(__file__))+'/'+"url.txt", 'w')
    f.close()
    print("생성된 url.txt 파일에 하나의 url을 입력하고 엔터를 반복한 후 다시 실행해주세요")
    sys.exit()

#=============================================
# stock_list = []
# time_list = []
# name_list = []

# # 재고유무 확인

# while True:
#     for url in urls:
#         response = requests.get(url)

#         if response.status_code == 200 :
#             html = response.text
#             soup = BeautifulSoup(html, 'html.parser')
#             title = soup.select_one('#stock-status')
#             text = title.get_text()
#             val = "품절" in text
#             if val:
#                 stock_list.append("재고 X")
#             else:
#                 stock_list.append("재고 O")
            
            
#             current = datetime.datetime.now()
#             current_time = current.strftime('%Y-%m-%d %H:%M') #current.replace(microsecond=0)
#             time_list.append(str(current_time))
#             name = soup.select_one('#name')
#             name_text = name.get_text()
#             name_list.append(name_text)

#         else:
#             print(response.status_code)
#             print("비정상적인 접근")

#     dataframe = pd.DataFrame({
#         'Time':time_list, 'name': name_list,'stock':stock_list
#     })

#     print(dataframe)

#     dataframe.to_excel('stocklist.xlsx')

#     time.sleep(60)
#============================== NEW
# Excel 구조잡기

wb = openpyxl.Workbook()
#첫리스트에 어떤 값이 올지 모름
for i in range(len(urls)):
    response = requests.get(urls[i])
    if response.status_code == 200:
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        name = soup.select_one('#name')
        name_text = name.get_text()
        print(name_text)
        # no_sheet_character = "\/*[]:?"

        # sheet이름에 특수문자 불가.
        #for x in range(len(no_sheet_character)):
        #    name_text = name_text.replace(no_sheet_character[x],"")
        #sheet 최대 문자열 길이 31개
        # if len(name_text) >30:
        #     name_text = name_text[:30]
        
        #sheet 이름에 제품명 붙이기
        # if i == 0 :
        #     wb['Sheet'].title = name_text
            
        # else:
        #     new_sheet = wb.create_sheet(name_text)
        if i == 0 :
            wb['Sheet'].title = str(i)
        else:
            new_sheet = wb.create_sheet(str(i))
        wb[str(i)].column_dimensions['A'].width = 25
        wb[str(i)].append(["날짜/시간", "재고"])
        wb[str(i)].cell(row = 1, column=3).value = name_text
    else:
        print("excel 구조 형성에서의 response status_code 오류")

wb.save(os.path.dirname(os.path.realpath(__file__))+'/'+"newstock.xlsx")

#색상지정
r_color = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
g_color = PatternFill(start_color='00e600', end_color='00e600', fill_type='solid')
i = 2
# 재고유무 확인
while True:

    for url_num in range(len(urls)):
        response = requests.get(urls[url_num])

        if response.status_code == 200 :
            #YYYY-MM-DD-HH-MM
            current = datetime.datetime.now()
            current_time = current.strftime('%Y-%m-%d %H:%M') #current.replace(microsecond=0)
            #time_list.append(str(current_time))
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            stock_status = soup.select_one('#stock-status')
            text = stock_status.get_text()
            val = "품절" in text
            if val:
                wb[str(url_num)].append([current_time, "재고 X"])
                wb[str(url_num)].cell(i,1).fill = r_color
                wb[str(url_num)].cell(i,2).fill = r_color
                #stock_list.append("재고 X")
            else:
                wb[str(url_num)].append([current_time, "재고 O"])
                wb[str(url_num)].cell(i,1).fill = g_color
                wb[str(url_num)].cell(i,2).fill = g_color
                #stock_list.append("재고 O")
        else:
            print(response.status_code)
            print("비정상적인 접근")
    i += 1
    wb.save(os.path.dirname(os.path.realpath(__file__))+'/'+"newstock.xlsx")
    print('저장완료')
    time.sleep(1)